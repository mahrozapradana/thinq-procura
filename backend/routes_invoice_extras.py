"""Invoice PDF export + bulk PO import + verified vendor pricelist lookup."""
from __future__ import annotations

import csv as _csv
import io as _io
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from auth_utils import get_current_active_user
from db_models import get_db, new_id, now_iso

router = APIRouter(prefix="/api")


def _fmt_num(v):
    try:
        return f"{float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(v or "-")


@router.get("/invoices/{iid}/pdf")
async def invoice_pdf(iid: str, user=Depends(get_current_active_user)):
    """Branded PDF of an invoice (accessible to admin/finance and to the owning vendor)."""
    db = get_db()
    inv = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not inv:
        raise HTTPException(404, "Invoice tidak ditemukan")
    if user["role"] == "vendor" and inv.get("vendor_id") != user.get("vendor_id"):
        raise HTTPException(403, "Not allowed")
    company = await db.company_settings.find_one({}, {"_id": 0}) or {}
    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.2*cm, bottomMargin=1.2*cm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=18, textColor=colors.HexColor("#0F172A"), spaceAfter=4)
    label = ParagraphStyle("label", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#64748B"), leading=10)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=10, leading=13)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#334155"), leading=10)

    elems = []
    # Header
    header_table = Table([[
        Paragraph(f"<b>{company.get('name') or 'PROCURA'}</b><br/><font size=8 color='#64748B'>{company.get('address') or 'E-Procurement Suite'}</font>", body),
        Paragraph("<para align='right'><font size=18 color='#0F172A'><b>INVOICE</b></font></para>", body),
    ]], colWidths=[10*cm, 7*cm])
    header_table.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP")]))
    elems.append(header_table)
    elems.append(Spacer(1, 8))

    # Meta grid
    meta_rows = [
        ["No Invoice", inv.get("invoice_number") or "-", "Tanggal", (inv.get("created_at") or "")[:10]],
        ["PO Referensi", inv.get("po_number") or "-", "Jatuh Tempo", inv.get("due_date") or "-"],
        ["Vendor", inv.get("vendor_name") or inv.get("vendor_id") or "-", "Status", (inv.get("status") or "").upper()],
        ["Currency", f"{inv.get('currency') or 'IDR'} · rate {inv.get('exchange_rate') or 1}", "Bonded", "Ya" if inv.get("is_bonded") else "-"],
    ]
    if inv.get("faktur_pajak_number"):
        meta_rows.append(["No. Faktur Pajak", inv["faktur_pajak_number"], "No. BAST", inv.get("bast_number") or "-"])
    meta = Table(meta_rows, colWidths=[3.5*cm, 5.5*cm, 3*cm, 5*cm])
    meta.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("TEXTCOLOR", (0,0), (0,-1), colors.HexColor("#64748B")),
        ("TEXTCOLOR", (2,0), (2,-1), colors.HexColor("#64748B")),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("TOPPADDING", (0,0), (-1,-1), 3),
    ]))
    elems.append(meta)
    elems.append(Spacer(1, 12))

    # Line items — prefer explicit line_items snapshot; fallback to po items
    lines = inv.get("line_items") or []
    if not lines:
        for i, it in enumerate(inv.get("items") or []):
            lines.append({
                "product_name": it.get("product_name"),
                "product_code": it.get("product_code"),
                "qty_billed": it.get("qty"),
                "unit_price": it.get("price"),
                "discount_amount": 0,
                "subtotal": it.get("subtotal") or (float(it.get("qty") or 0) * float(it.get("price") or 0)),
            })
    li_rows = [["#", "Produk", "Qty", "Harga", "Diskon", "Subtotal"]]
    for i, li in enumerate(lines, start=1):
        li_rows.append([
            str(i),
            f"{li.get('product_name') or '-'}\n{li.get('product_code') or ''}",
            _fmt_num(li.get("qty_billed")),
            _fmt_num(li.get("unit_price")),
            _fmt_num(li.get("discount_amount")),
            _fmt_num(li.get("subtotal")),
        ])
    li_table = Table(li_rows, colWidths=[0.8*cm, 7.2*cm, 2*cm, 2.5*cm, 2*cm, 2.5*cm], repeatRows=1)
    li_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
        ("ALIGN", (2,1), (-1,-1), "RIGHT"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    elems.append(li_table)
    elems.append(Spacer(1, 8))

    # Tax breakdown (optional)
    tbs = inv.get("tax_breakdown") or []
    if tbs:
        tb_rows = [["Kode", "Nama", "Rate", "Base", "Jumlah"]]
        for t in tbs:
            amount = float(t.get("amount") or 0)
            sign = "-" if t.get("tax_type") == "withholding" else ""
            tb_rows.append([
                t.get("code") or "-",
                t.get("name") or "-",
                f"{t.get('rate') or 0}%",
                _fmt_num(t.get("base")),
                f"{sign}{_fmt_num(amount)}",
            ])
        tb_table = Table(tb_rows, colWidths=[2.5*cm, 6*cm, 2*cm, 3*cm, 3.5*cm], repeatRows=1)
        tb_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#334155")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
            ("ALIGN", (2,1), (-1,-1), "RIGHT"),
        ]))
        elems.append(Paragraph("<b>Rincian Pajak</b>", body))
        elems.append(Spacer(1, 4))
        elems.append(tb_table)
        elems.append(Spacer(1, 8))

    # Totals
    total_rows = [
        ["Subtotal (Untaxed)", _fmt_num(inv.get("untaxed_amount"))],
        ["Total Pajak", _fmt_num(inv.get("amount_tax"))],
        ["GRAND TOTAL", _fmt_num(inv.get("amount_total") or inv.get("amount"))],
    ]
    if (inv.get("vendor_reply") or {}).get("totals", {}).get("discount_amount"):
        total_rows.insert(1, ["Diskon Vendor", f"- {_fmt_num(inv['vendor_reply']['totals']['discount_amount'])}"])
    tt = Table(total_rows, colWidths=[13*cm, 4*cm])
    tt.setStyle(TableStyle([
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-2), 9),
        ("FONTSIZE", (0,-1), (-1,-1), 11),
        ("ALIGN", (1,0), (1,-1), "RIGHT"),
        ("LINEABOVE", (0,-1), (-1,-1), 1, colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0,-1), (-1,-1), colors.HexColor("#0F172A")),
        ("TOPPADDING", (0,-1), (-1,-1), 6),
    ]))
    elems.append(tt)
    elems.append(Spacer(1, 12))

    if inv.get("notes"):
        elems.append(Paragraph("<b>Catatan:</b>", body))
        elems.append(Paragraph(inv["notes"], small))
        elems.append(Spacer(1, 8))

    docs_lines = []
    if inv.get("faktur_pajak_url"):
        docs_lines.append(f"• Faktur Pajak: {inv.get('faktur_pajak_filename') or inv.get('faktur_pajak_url')}")
    if inv.get("bast_url"):
        docs_lines.append(f"• BAST: {inv.get('bast_filename') or inv.get('bast_url')}")
    for a in (inv.get("attachments") or []):
        docs_lines.append(f"• {a.get('filename') or a.get('url')}")
    if docs_lines:
        elems.append(Paragraph("<b>Dokumen Terlampir</b>", body))
        for line in docs_lines:
            elems.append(Paragraph(line, small))

    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={
        "Content-Disposition": f'attachment; filename="{inv.get("invoice_number") or "invoice"}.pdf"'
    })


# ---------- Bulk PO Import ----------
@router.post("/pos/bulk-import")
async def bulk_import_po(file: UploadFile = File(...), user=Depends(get_current_active_user)):
    """Bulk create POs from CSV/XLSX. Rows grouped by vendor_code.

    Columns required: vendor_code, product_code, qty, price.
    Optional: po_type (LOCAL|BONDED), currency (IDR default), delivery_date, notes.
    Each unique (vendor_code, po_type, currency) forms one PO.
    """
    if user["role"] not in ("admin", "procurement"):
        raise HTTPException(403, "Not allowed")
    db = get_db()
    content = await file.read()
    rows: list[dict] = []
    ext = (file.filename or "").lower()
    try:
        if ext.endswith(".xlsx") or ext.endswith(".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(_io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not any(r):
                    continue
                rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
        else:
            reader = _csv.DictReader(_io.StringIO(content.decode("utf-8-sig")))
            rows = [{k.strip().lower(): v for k, v in row.items() if k} for row in reader]
    except Exception as e:
        raise HTTPException(400, f"Gagal parsing file: {e}")

    vendors_by_code = {v.get("code"): v async for v in db.vendors.find({"status": "approved"}, {"_id": 0})}
    prods_by_code = {p.get("code"): p async for p in db.products.find({}, {"_id": 0})}

    groups: dict[tuple, dict] = {}
    errors: list[dict] = []
    for i, r in enumerate(rows, start=2):
        vcode = str(r.get("vendor_code") or "").strip()
        pcode = str(r.get("product_code") or "").strip()
        v = vendors_by_code.get(vcode)
        p = prods_by_code.get(pcode)
        if not v:
            errors.append({"row": i, "error": f"vendor_code '{vcode}' tidak ditemukan"})
            continue
        if not p:
            errors.append({"row": i, "error": f"product_code '{pcode}' tidak ditemukan"})
            continue
        try:
            qty = float(r.get("qty") or 0)
            price = float(r.get("price") or 0)
        except Exception:
            errors.append({"row": i, "error": "qty/price tidak valid"})
            continue
        if qty <= 0 or price < 0:
            errors.append({"row": i, "error": "qty harus > 0"})
            continue
        po_type = str(r.get("po_type") or "LOCAL").upper()
        currency = str(r.get("currency") or "IDR").upper()
        key = (v["id"], po_type, currency)
        g = groups.setdefault(key, {
            "vendor_id": v["id"],
            "vendor_name": v.get("company_name"),
            "po_type": po_type,
            "currency": currency,
            "delivery_date": str(r.get("delivery_date") or "") or None,
            "notes": str(r.get("notes") or "") or None,
            "items": [],
        })
        g["items"].append({
            "product_id": p["id"],
            "product_name": p.get("name"),
            "product_code": p.get("code"),
            "qty": qty,
            "price": price,
            "subtotal": qty * price,
        })

    created: list[dict] = []
    from db_models import get_db as _  # noqa
    # Import here to avoid a circular import at module load
    for key, g in groups.items():
        total = sum(it["subtotal"] for it in g["items"])
        po_doc = {
            "id": new_id(),
            "po_number": f"PO-BULK-{new_id()[:8].upper()}",
            "vendor_id": g["vendor_id"],
            "vendor_name": g["vendor_name"],
            "po_type": g["po_type"],
            "currency": g["currency"],
            "exchange_rate": 1.0,
            "items": g["items"],
            "total": total,
            "untaxed_amount": total,
            "amount_total": total,
            "amount_tax": 0,
            "tax_ids": [],
            "tax_breakdown": [],
            "status": "draft",
            "invoice_status": "pending",
            "shipping_status": "pending",
            "delivery_date": g["delivery_date"],
            "notes": g["notes"],
            "created_by": user["id"],
            "created_at": now_iso(),
            "source": "bulk_import",
        }
        await db.pos.insert_one(po_doc)
        created.append({"id": po_doc["id"], "po_number": po_doc["po_number"], "items": len(g["items"]), "total": total})
    return {"ok": True, "created_count": len(created), "created": created, "errors": errors, "total_rows": len(rows)}


# ---------- Cheapest verified vendor pricelist ----------
@router.get("/pricelists/cheapest")
async def cheapest_pricelist(product_id: str, only_verified: bool = True, user=Depends(get_current_active_user)):
    """Return the cheapest vendor pricelist for a product. Default excludes unverified rows."""
    db = get_db()
    q: dict = {"product_id": product_id}
    if only_verified:
        q["verified"] = True
    rows = await db.vendor_pricelists.find(q, {"_id": 0}).sort("price", 1).to_list(100)
    return {"product_id": product_id, "count": len(rows), "cheapest": rows[0] if rows else None, "top3": rows[:3]}
