"""SPT-1111 Excel template export (Formulir 1111 B2 - Pajak Masukan)."""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from auth_utils import get_current_active_user
from db_models import get_db

router = APIRouter(prefix="/api/reports")


@router.get("/spt-1111.xlsx")
async def spt_1111_xlsx(
    year: int,
    month: int,
    user=Depends(get_current_active_user),
):
    """Generate SPT PPN 1111 Form B2 (Pajak Masukan yang Dapat Dikreditkan)."""
    if user["role"] not in ("admin", "procurement", "finance"):
        raise HTTPException(403, "Not allowed")
    db = get_db()
    # Aggregate invoices for the month
    start = f"{year:04d}-{month:02d}-01"
    nm = month + 1
    ny = year + 1 if nm > 12 else year
    nm = 1 if nm > 12 else nm
    end = f"{ny:04d}-{nm:02d}-01"
    # Use invoices (which represent tax invoices from vendor) OR fallback to POs
    invoices = await db.invoices.find({"created_at": {"$gte": start, "$lt": end}}, {"_id": 0}).sort("created_at", 1).to_list(5000)
    pos_map = {p["id"]: p for p in await db.pos.find({}, {"_id": 0}).to_list(5000)}
    vendors = {v["id"]: v for v in await db.vendors.find({}, {"_id": 0}).to_list(2000)}

    wb = Workbook()

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    title_font = Font(bold=True, size=11)

    # === Sheet 1: FORMULIR 1111 B2 — Pajak Masukan ===
    ws = wb.active
    ws.title = "1111 B2 Pajak Masukan"

    # Header block (DJP-style)
    ws["A1"] = "FORMULIR 1111 B2"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:J1")
    ws["A2"] = "DAFTAR PAJAK MASUKAN YANG DAPAT DIKREDITKAN"
    ws["A2"].font = Font(bold=True, size=11)
    ws.merge_cells("A2:J2")
    ws["A3"] = "DAN/ATAU PPn BM"
    ws["A3"].font = Font(bold=True, size=10)
    ws.merge_cells("A3:J3")

    ws["A5"] = "NAMA PKP"
    ws["C5"] = "(diisi Nama PKP)"
    ws["A6"] = "NPWP"
    ws["C6"] = "(diisi NPWP PKP)"
    ws["A7"] = "MASA"
    ws["C7"] = f"{month:02d} - {year}"
    ws["C7"].font = Font(bold=True)

    for cell in ("A5", "A6", "A7"):
        ws[cell].font = Font(bold=True)

    # Column headers row 9-10
    headers_top = [
        ("NO.", None), ("NAMA PENJUAL BKP/BJP", None), ("NPWP", None),
        ("FAKTUR PAJAK / DOKUMEN TERTENTU", 3),
        ("DPP (Rupiah)", None), ("PPN (Rupiah)", None), ("PPnBM (Rupiah)", None), ("KET.", None),
    ]
    col = 1
    for label, colspan in headers_top:
        cell = ws.cell(row=9, column=col, value=label)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if colspan:
            ws.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col+colspan-1)
            col += colspan
        else:
            ws.merge_cells(start_row=9, start_column=col, end_row=10, end_column=col)
            col += 1
    # Sub-headers under Faktur
    for i, sub in enumerate(["KODE & NO. SERI", "TANGGAL", "REF"]):
        cell = ws.cell(row=10, column=4+i, value=sub)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data rows
    row = 11
    total_dpp = 0.0
    total_ppn = 0.0
    seq = 0
    for inv in invoices:
        po = pos_map.get(inv.get("po_id"), {})
        vendor = vendors.get(inv.get("vendor_id"), {})
        untaxed = float(po.get("untaxed_amount") or po.get("total") or inv.get("amount") or 0)
        # Sum only sales taxes (PPN) from breakdown
        ppn = sum(float(t.get("amount") or 0) for t in (po.get("tax_breakdown") or []) if t.get("tax_type") != "withholding")
        if not ppn:
            ppn = float(po.get("amount_tax") or 0)
        if ppn <= 0 and untaxed <= 0:
            continue
        seq += 1
        vals = [
            seq,
            vendor.get("company_name") or vendor.get("name") or "-",
            (vendor.get("npwp") or "").replace(".", "").replace("-", "")[:16],
            inv.get("invoice_number") or "-",
            (inv.get("created_at") or "")[:10],
            po.get("po_number") or "-",
            untaxed,
            ppn,
            0.0,
            "01",  # kode transaksi standar
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
            if c in (7, 8, 9):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
        total_dpp += untaxed
        total_ppn += ppn
        row += 1

    # Fallback: if no invoices, use POs in period so form still populates
    if seq == 0:
        for pid, po in pos_map.items():
            if not ((po.get("created_at") or "") >= start and (po.get("created_at") or "") < end):
                continue
            if po.get("status") not in ("approved", "sent", "partial", "completed"):
                continue
            vendor = vendors.get(po.get("vendor_id"), {})
            untaxed = float(po.get("untaxed_amount") or po.get("total") or 0)
            ppn = sum(float(t.get("amount") or 0) for t in (po.get("tax_breakdown") or []) if t.get("tax_type") != "withholding") or float(po.get("amount_tax") or 0)
            if ppn <= 0:
                continue
            seq += 1
            vals = [seq, vendor.get("company_name") or "-", (vendor.get("npwp") or "").replace(".", "").replace("-", "")[:16],
                    po.get("po_number") or "-", (po.get("created_at") or "")[:10], po.get("po_number") or "-",
                    untaxed, ppn, 0.0, "01"]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = border
                if c in (7, 8, 9):
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
            total_dpp += untaxed
            total_ppn += ppn
            row += 1

    # Total row
    ws.cell(row=row, column=1, value="JUMLAH").font = Font(bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    for c, v in enumerate([total_dpp, total_ppn, 0.0], start=7):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.number_format = "#,##0"
        cell.border = border
    row += 2
    ws.cell(row=row, column=1, value="Digenerasi Procura E-Procurement Suite pada " + datetime.now().strftime("%d %b %Y %H:%M")).font = Font(italic=True, size=9, color="64748B")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    # Column widths
    widths = [6, 32, 18, 24, 12, 14, 16, 16, 16, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # === Sheet 2: INDUK Ringkasan ===
    ws2 = wb.create_sheet("Induk Ringkasan")
    ws2["A1"] = f"SPT MASA PPN — Periode {month:02d}/{year}"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:D1")
    rows_summary = [
        ("A. Pajak Keluaran", "-"),
        ("B.1 Pajak Masukan dari Faktur Pajak", f"{total_ppn:,.0f}".replace(",", ".")),
        ("B.2 Pajak Masukan yang Dapat Dikreditkan", f"{total_ppn:,.0f}".replace(",", ".")),
        ("Jumlah PPN Terutang / Lebih Bayar", f"({total_ppn:,.0f})".replace(",", ".")),
    ]
    for i, (label, val) in enumerate(rows_summary, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=(i == 6))
        ws2.cell(row=i, column=3, value=val).alignment = Alignment(horizontal="right")
    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["C"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"SPT_1111_{year}-{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
