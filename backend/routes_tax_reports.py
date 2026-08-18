"""Monthly tax report (SPT-ready) — Excel + PDF export."""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from auth_utils import get_current_active_user
from db_models import get_db

router = APIRouter(prefix="/api/reports")


async def _aggregate_taxes(db, year: int, month: Optional[int]):
    """Aggregate tax data from POs in the given period."""
    if month:
        start = f"{year:04d}-{month:02d}-01"
        nm = month + 1
        ny = year + 1 if nm > 12 else year
        nm = 1 if nm > 12 else nm
        end = f"{ny:04d}-{nm:02d}-01"
    else:
        start = f"{year:04d}-01-01"
        end = f"{year+1:04d}-01-01"
    pos = await db.pos.find({
        "created_at": {"$gte": start, "$lt": end},
        "status": {"$in": ["approved", "sent", "partial", "completed"]},
    }, {"_id": 0}).sort("created_at", 1).to_list(5000)
    vendors = {v["id"]: v.get("company_name") for v in await db.vendors.find({}, {"_id": 0}).to_list(2000)}
    rows = []
    totals = {"untaxed": 0.0, "sales_tax": 0.0, "withholding": 0.0, "grand": 0.0}
    tax_summary: dict[str, dict] = {}  # by tax code
    for p in pos:
        untaxed = float(p.get("untaxed_amount") or p.get("total") or 0)
        totals["untaxed"] += untaxed
        po_sales, po_wh = 0.0, 0.0
        for tx in (p.get("tax_breakdown") or []):
            amt = float(tx.get("amount") or 0)
            code = tx.get("code") or tx.get("name") or "-"
            tax_summary.setdefault(code, {"code": code, "name": tx.get("name"), "rate": tx.get("rate"), "tax_type": tx.get("tax_type"), "base_total": 0.0, "amount_total": 0.0, "po_count": 0})
            tax_summary[code]["base_total"] += float(tx.get("base") or untaxed)
            tax_summary[code]["amount_total"] += amt
            tax_summary[code]["po_count"] += 1
            if tx.get("tax_type") == "withholding":
                po_wh += amt
                totals["withholding"] += amt
            else:
                po_sales += amt
                totals["sales_tax"] += amt
        # legacy tax_percent path
        if not p.get("tax_breakdown") and float(p.get("amount_tax") or 0):
            amt = float(p["amount_tax"])
            po_sales += amt
            totals["sales_tax"] += amt
            key = f"PPN{int(p.get('tax_percent') or 0)}"
            tax_summary.setdefault(key, {"code": key, "name": f"PPN {p.get('tax_percent') or 0}%", "rate": p.get("tax_percent"), "tax_type": "sales", "base_total": 0.0, "amount_total": 0.0, "po_count": 0})
            tax_summary[key]["base_total"] += untaxed
            tax_summary[key]["amount_total"] += amt
            tax_summary[key]["po_count"] += 1
        rows.append({
            "po_number": p.get("po_number"),
            "date": (p.get("created_at") or "")[:10],
            "vendor": vendors.get(p.get("vendor_id"), "-"),
            "vendor_npwp": "",  # populate later
            "untaxed": untaxed,
            "sales_tax": po_sales,
            "withholding": po_wh,
            "grand_total": float(p.get("amount_total") or (untaxed + po_sales - po_wh)),
            "tax_codes": ", ".join(t.get("code", "-") for t in (p.get("tax_breakdown") or [])) or (f"PPN{int(p.get('tax_percent') or 0)}" if p.get("tax_percent") else "-"),
        })
    totals["grand"] = totals["untaxed"] + totals["sales_tax"] - totals["withholding"]
    return rows, list(tax_summary.values()), totals


@router.get("/taxes.xlsx")
async def taxes_xlsx(
    year: int,
    month: Optional[int] = None,
    user=Depends(get_current_active_user),
):
    if user["role"] not in ("admin", "procurement", "finance"):
        raise HTTPException(403, "Not allowed")
    db = get_db()
    rows, summary, totals = await _aggregate_taxes(db, year, month)

    wb = Workbook()
    thin = Side(border_style="thin", color="94A3B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    label_fill = PatternFill("solid", fgColor="F1F5F9")

    # Sheet 1: Detail per PO
    ws1 = wb.active
    ws1.title = "Detail PO"
    period_label = f"{year}-{month:02d}" if month else f"{year}"
    ws1["A1"] = f"Laporan Pajak per PO — Periode {period_label}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:I1")
    ws1["A2"] = f"Digenerasi: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A2"].font = Font(italic=True, color="64748B", size=9)
    ws1.merge_cells("A2:I2")

    headers = ["No PO", "Tanggal", "Vendor", "NPWP", "Kode Pajak", "DPP (Untaxed)", "PPN Keluar", "Potongan PPh", "Grand Total"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for i, r in enumerate(rows, start=5):
        vals = [r["po_number"], r["date"], r["vendor"], r["vendor_npwp"], r["tax_codes"], r["untaxed"], r["sales_tax"], r["withholding"], r["grand_total"]]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=c, value=v)
            cell.border = border
            if c >= 6:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
    # totals row
    tr = len(rows) + 5
    ws1.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    ws1.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=5)
    for c, v in enumerate([totals["untaxed"], totals["sales_tax"], totals["withholding"], totals["grand"]], start=6):
        cell = ws1.cell(row=tr, column=c, value=v)
        cell.font = Font(bold=True)
        cell.fill = label_fill
        cell.number_format = "#,##0"
        cell.border = border
    # Column widths
    widths = [18, 12, 32, 18, 20, 16, 16, 16, 18]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Sheet 2: Ringkasan per Jenis Pajak
    ws2 = wb.create_sheet("Ringkasan Pajak")
    ws2["A1"] = f"Ringkasan per Jenis Pajak — {period_label}"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:F1")
    headers2 = ["Kode", "Nama", "Tipe", "Rate %", "Base (DPP)", "Total Amount", "Jumlah PO"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for i, s in enumerate(summary, start=4):
        vals = [s["code"], s["name"], s["tax_type"], s["rate"], s["base_total"], s["amount_total"], s["po_count"]]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.border = border
            if c in (5, 6):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
    widths2 = [14, 28, 14, 10, 18, 18, 12]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"tax_report_{period_label}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/taxes/summary")
async def taxes_summary(
    year: int,
    month: Optional[int] = None,
    user=Depends(get_current_active_user),
):
    if user["role"] not in ("admin", "procurement", "finance"):
        raise HTTPException(403, "Not allowed")
    db = get_db()
    rows, summary, totals = await _aggregate_taxes(db, year, month)
    return {"totals": totals, "summary": summary, "row_count": len(rows), "period": f"{year}-{month:02d}" if month else str(year)}
