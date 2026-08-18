"""Vendor portal: self-registration + vendor-facing endpoints."""
from __future__ import annotations

from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel, EmailStr

from auth_utils import get_current_active_user
from db_models import get_db, new_id, now_iso, clean, gen_number

router = APIRouter(prefix="/api")


class VendorRegisterIn(BaseModel):
    company_name: str
    name: str  # contact person
    email: EmailStr
    phone: Optional[str] = None
    address: Optional[str] = None
    npwp: Optional[str] = None
    is_importer: bool = False
    categories: List[str] = []
    bank_account: Optional[str] = None
    description: Optional[str] = None


@router.post("/vendor/register")
async def vendor_register(payload: VendorRegisterIn):
    """Public vendor self-registration."""
    db = get_db()
    email = payload.email.lower()
    existing = await db.vendors.find_one({"email": email})
    if existing:
        raise HTTPException(400, "Email vendor sudah terdaftar")
    doc = {
        "id": new_id(),
        "status": "pending_approval",
        "user_id": None,
        "created_at": now_iso(),
        **payload.model_dump(),
        "email": email,
    }
    await db.vendors.insert_one(doc)
    return {"ok": True, "vendor_id": doc["id"], "message": "Pendaftaran diterima. Tunggu persetujuan tim procurement."}


# ---------- Vendor portal endpoints ----------
def _require_vendor(user: dict) -> str:
    if user["role"] != "vendor" or not user.get("vendor_id"):
        raise HTTPException(403, "Vendor portal only")
    return user["vendor_id"]


@router.get("/vendor-portal/profile")
async def get_vendor_profile(user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    v = await db.vendors.find_one({"id": vid}, {"_id": 0})
    return v


class ProfileUpdateIn(BaseModel):
    phone: Optional[str] = None
    address: Optional[str] = None
    npwp: Optional[str] = None
    bank_account: Optional[str] = None
    description: Optional[str] = None
    categories: Optional[List[str]] = None


@router.put("/vendor-portal/profile")
async def update_vendor_profile(payload: ProfileUpdateIn, user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    upd = {k: v for k, v in payload.model_dump().items() if v is not None}
    await db.vendors.update_one({"id": vid}, {"$set": upd})
    return await db.vendors.find_one({"id": vid}, {"_id": 0})


@router.get("/vendor-portal/tenders")
async def vendor_tenders(user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    # Only OPEN tenders OR tenders where this vendor has participated (bid/decline) OR was invited
    q = {
        "$or": [
            {"status": "open", "$or": [{"invited_vendor_ids": vid}, {"invited_vendor_ids": {"$in": [None, []]}}, {"invited_vendor_ids": {"$exists": False}}]},
            {"invited_vendor_ids": vid},
            {"bids.vendor_id": vid},
            {"awarded_vendor_id": vid},
        ],
    }
    return await db.tenders.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.get("/vendor-portal/rfqs")
async def vendor_rfqs(user=Depends(get_current_active_user)):
    """List of POs directed to this vendor that are still draft or pending approval (RFQ / pre-PO)."""
    vid = _require_vendor(user)
    db = get_db()
    q = {"vendor_id": vid, "status": {"$in": ["draft", "pending_approval"]}}
    if user.get("is_pic"):
        q["assigned_pic_id"] = user["id"]
    return await db.pos.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.get("/vendor-portal/pos/{pid}")
async def vendor_po_detail(pid: str, user=Depends(get_current_active_user)):
    """Read-only PO detail for vendor including tax breakdown."""
    vid = _require_vendor(user)
    db = get_db()
    po = await db.pos.find_one({"id": pid, "vendor_id": vid}, {"_id": 0})
    if not po:
        raise HTTPException(404, "PO tidak ditemukan atau bukan milik vendor Anda")
    if user.get("is_pic") and po.get("assigned_pic_id") and po["assigned_pic_id"] != user["id"]:
        raise HTTPException(403, "PO ini tidak di-assign ke PIC Anda")
    return po


@router.get("/vendor-portal/unread-counts")
async def vendor_unread_counts(user=Depends(get_current_active_user)):
    """Sidebar badge counts: RFQ (draft/pending POs), PO (unacknowledged), Invoice (outstanding), Tender (open, not-bid)."""
    vid = _require_vendor(user)
    db = get_db()
    q_scope = {"vendor_id": vid}
    if user.get("is_pic"):
        q_scope["assigned_pic_id"] = user["id"]
    rfq = await db.pos.count_documents({**q_scope, "status": {"$in": ["draft", "pending_approval"]}, "vendor_reply": {"$exists": False}})
    po_new = await db.pos.count_documents({**q_scope, "status": {"$in": ["approved", "sent"]}, "vendor_acknowledged": {"$ne": True}})
    invoice_out = await db.invoices.count_documents({"vendor_id": vid, "status": {"$in": ["outstanding", "pending"]}})
    tender = await db.tenders.count_documents({
        "status": "open",
        "$or": [{"invited_vendor_ids": vid}, {"invited_vendor_ids": {"$in": [None, []]}}, {"invited_vendor_ids": {"$exists": False}}],
        "bids": {"$not": {"$elemMatch": {"vendor_id": vid, "status": "submitted"}}},
    })
    return {"rfq": rfq, "po": po_new, "invoice": invoice_out, "tender": tender}


@router.get("/internal/unread-counts")
async def internal_unread_counts(user=Depends(get_current_active_user)):
    """Sidebar badge counts for internal roles (buyer/admin/procurement/finance/warehouse)."""
    if user.get("role") == "vendor":
        raise HTTPException(403, "Internal only")
    db = get_db()
    pr_pending = await db.prs.count_documents({"status": "pending_approval"})
    po_pending = await db.pos.count_documents({"status": "pending_approval"})
    tender_open = await db.tenders.count_documents({"status": "open"})
    vendor_pending = await db.vendors.count_documents({"status": "pending"})
    invoice_outstanding = await db.invoices.count_documents({"status": {"$in": ["outstanding", "pending"]}})
    customs_draft = await db.customs_docs.count_documents({"status": "draft"}) if "customs_docs" in await db.list_collection_names() else 0
    receipt_pending = await db.pos.count_documents({"status": "sent", "shipping_status": {"$in": ["waiting_delivery", "in_transit"]}})
    return {
        "pr": pr_pending,
        "po": po_pending,
        "tender": tender_open,
        "vendors": vendor_pending,
        "invoices": invoice_outstanding,
        "customs": customs_draft,
        "receipts": receipt_pending,
    }


@router.post("/vendor-portal/pos/{pid}/acknowledge")
async def vendor_acknowledge_po(pid: str, user=Depends(get_current_active_user)):
    """Vendor confirms receipt of an approved/sent PO."""
    vid = _require_vendor(user)
    db = get_db()
    po = await db.pos.find_one({"id": pid, "vendor_id": vid})
    if not po:
        raise HTTPException(404, "PO tidak ditemukan")
    if po.get("status") not in ("approved", "sent"):
        raise HTTPException(400, "Hanya PO status approved/sent yang bisa diakui")
    if po.get("vendor_acknowledged"):
        raise HTTPException(400, "PO ini sudah diakui sebelumnya")
    await db.pos.update_one(
        {"id": pid},
        {"$set": {
            "vendor_acknowledged": True,
            "vendor_acknowledged_at": now_iso(),
            "vendor_acknowledged_by": user["id"],
            "vendor_acknowledged_by_name": user.get("name"),
        }},
    )
    return {"ok": True, "acknowledged_at": now_iso()}


class RFQReplyItem(BaseModel):
    item_index: int
    price: Optional[float] = None  # counter price per unit (pre-discount)
    discount_type: Optional[str] = None  # 'percent' | 'amount' | None
    discount_value: Optional[float] = 0
    notes: Optional[str] = None


class RFQReplyIn(BaseModel):
    items: List[RFQReplyItem] = []
    overall_notes: Optional[str] = None
    delivery_days: Optional[int] = None
    can_fulfill: bool = True  # false = decline outright


def _apply_discount(price: float, qty: float, discount_type: Optional[str], discount_value: Optional[float]):
    subtotal = float(price) * float(qty)
    dv = float(discount_value or 0)
    if not dv or not discount_type:
        return {"subtotal_before": subtotal, "discount_amount": 0.0, "subtotal_after": subtotal}
    if discount_type == "percent":
        amt = subtotal * (dv / 100.0)
    else:  # 'amount' → rupiah per unit
        amt = dv * float(qty)
    amt = max(0.0, min(amt, subtotal))
    return {"subtotal_before": subtotal, "discount_amount": amt, "subtotal_after": subtotal - amt}


@router.post("/vendor-portal/rfqs/{pid}/reply")
async def vendor_rfq_reply(pid: str, payload: RFQReplyIn, user=Depends(get_current_active_user)):
    """Vendor sends counter/confirm prices (with optional per-item discount)."""
    vid = _require_vendor(user)
    db = get_db()
    po = await db.pos.find_one({"id": pid, "vendor_id": vid})
    if not po:
        raise HTTPException(404, "RFQ tidak ditemukan")
    if po.get("status") not in ("draft", "pending_approval"):
        raise HTTPException(400, "RFQ sudah tidak dapat direspons (status berubah)")
    # Snapshot items w/ discount computation for buyer preview
    items_out: list[dict] = []
    total_before = 0.0
    total_discount = 0.0
    total_after = 0.0
    for it in payload.items:
        po_it = (po.get("items") or [])[it.item_index] if it.item_index < len(po.get("items") or []) else {}
        qty = float(po_it.get("qty") or 0)
        price = float(it.price if it.price is not None else po_it.get("price") or 0)
        disc = _apply_discount(price, qty, it.discount_type, it.discount_value)
        total_before += disc["subtotal_before"]
        total_discount += disc["discount_amount"]
        total_after += disc["subtotal_after"]
        items_out.append({
            **it.model_dump(),
            "qty": qty,
            "subtotal_before": disc["subtotal_before"],
            "discount_amount": disc["discount_amount"],
            "subtotal_after": disc["subtotal_after"],
        })
    reply_doc = {
        "vendor_id": vid,
        "vendor_name": user.get("name"),
        "can_fulfill": payload.can_fulfill,
        "delivery_days": payload.delivery_days,
        "overall_notes": payload.overall_notes,
        "items": items_out,
        "totals": {
            "before_discount": total_before,
            "discount_amount": total_discount,
            "after_discount": total_after,
        },
        "submitted_at": now_iso(),
    }
    await db.pos.update_one({"id": pid}, {"$set": {"vendor_reply": reply_doc}})
    # Notify buyer via bell notification
    try:
        from routes_notifications import create_notification, notify_role
        if po.get("created_by"):
            await create_notification(
                po["created_by"],
                "rfq_reply",
                f"Vendor membalas RFQ {po.get('po_number')}",
                f"{user.get('name')} kirim {'penawaran counter' if payload.can_fulfill else 'penolakan'} — cek detail sekarang.",
                f"/po?open={pid}",
                {"po_id": pid, "po_number": po.get("po_number")},
            )
        # Also notify all procurement users
        await notify_role(
            "procurement", "rfq_reply",
            f"Vendor membalas RFQ {po.get('po_number')}",
            f"{user.get('name')} kirim balasan — perlu ditinjau",
            f"/po?open={pid}",
        )
    except Exception:
        pass
    return {"ok": True, "reply": reply_doc}


class BidItemIn(BaseModel):
    item_index: int
    can_fulfill: bool = True
    qty_offered: Optional[float] = None  # bila vendor hanya sanggup sebagian
    price: Optional[float] = None  # harga per unit
    notes: Optional[str] = None


class BidAttachmentIn(BaseModel):
    url: str
    filename: str
    size: Optional[int] = None
    content_type: Optional[str] = None


class BidIn(BaseModel):
    price: float  # total
    delivery_days: Optional[int] = None
    notes: Optional[str] = None
    items: List[BidItemIn] = []
    attachments: List[BidAttachmentIn] = []
    is_draft: bool = False


@router.get("/vendor-portal/tenders/{tid}")
async def vendor_tender_detail(tid: str, user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    t = await db.tenders.find_one({"id": tid}, {"_id": 0})
    if not t:
        raise HTTPException(404, "Tender not found")
    # Only allow if vendor is invited or tender is public/open
    invited = t.get("invited_vendor_ids") or []
    if invited and vid not in invited and not any(b.get("vendor_id") == vid for b in (t.get("bids") or [])):
        raise HTTPException(403, "Tender bukan untuk Anda")
    # Hide competitor bids
    my_bid = next((b for b in (t.get("bids") or []) if b.get("vendor_id") == vid), None)
    t.pop("bids", None)
    t["my_bid"] = my_bid
    # Add sealed indicator for vendor UI
    if t.get("is_sealed"):
        t["sealed_revealed"] = bool(t.get("sealed_revealed_at"))
    return t


@router.post("/vendor-portal/tenders/{tid}/bid")
async def submit_bid(tid: str, payload: BidIn, user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    t = await db.tenders.find_one({"id": tid})
    if not t:
        raise HTTPException(404, "Tender not found")
    if t.get("status") != "open":
        raise HTTPException(400, "Tender not open")
    # Deadline lock — reject once deadline passed (drafts also blocked to avoid last-minute prep loopholes)
    if t.get("deadline"):
        try:
            from datetime import datetime, timezone
            dl = t["deadline"]
            dl_dt = datetime.fromisoformat(dl.replace("Z", "+00:00")) if "T" in dl else datetime.fromisoformat(dl + "T23:59:59+00:00")
            if dl_dt.tzinfo is None:
                dl_dt = dl_dt.replace(tzinfo=timezone.utc)
            if datetime.now(timezone.utc) > dl_dt:
                raise HTTPException(400, "Deadline tender sudah lewat — bid tidak dapat disubmit")
        except HTTPException:
            raise
        except Exception:
            pass  # malformed deadline → allow
    bids = t.get("bids", [])
    prev = next((b for b in bids if b.get("vendor_id") == vid), None)
    history = list(prev.get("history") or []) if prev else []
    if prev and prev.get("status") in ("draft", "submitted"):
        history.append({
            "price": prev.get("price"),
            "delivery_days": prev.get("delivery_days"),
            "notes": prev.get("notes"),
            "status": prev.get("status"),
            "submitted_at": prev.get("submitted_at"),
            "items": prev.get("items") or [],
        })
        # cap history to last 20 versions
        history = history[-20:]
    bids = [b for b in bids if b["vendor_id"] != vid]  # remove old bid
    bids.append({
        "vendor_id": vid,
        "vendor_name": user["name"],
        "price": payload.price,
        "delivery_days": payload.delivery_days,
        "notes": payload.notes,
        "items": [i.model_dump() for i in payload.items],
        "attachments": [a.model_dump() for a in payload.attachments],
        "status": "draft" if payload.is_draft else "submitted",
        "submitted_at": now_iso(),
        "history": history,
    })
    await db.tenders.update_one({"id": tid}, {"$set": {"bids": bids}})
    return {"ok": True, "is_draft": payload.is_draft, "history_count": len(history)}


# ---------- Vendor pricelists ----------
class VendorPricelistIn(BaseModel):
    product_id: str
    price: float
    currency: str = "IDR"
    min_qty: Optional[float] = 1
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None
    notes: Optional[str] = None
    file_url: Optional[str] = None
    filename: Optional[str] = None


@router.get("/vendor-portal/pricelists")
async def list_vendor_pricelists(user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    rows = await db.vendor_pricelists.find({"vendor_id": vid}, {"_id": 0}).sort("created_at", -1).to_list(500)
    prods = {p["id"]: p async for p in db.products.find({}, {"_id": 0, "id": 1, "name": 1, "code": 1})}
    for r in rows:
        p = prods.get(r.get("product_id"))
        if p:
            r["product_name"] = p.get("name")
            r["product_code"] = p.get("code")
    return rows


@router.post("/vendor-portal/pricelists")
async def create_vendor_pricelist(payload: VendorPricelistIn, user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    prod = await db.products.find_one({"id": payload.product_id}, {"_id": 0, "name": 1})
    if not prod:
        raise HTTPException(404, "Produk tidak ditemukan")
    doc = {
        "id": new_id(),
        "vendor_id": vid,
        "vendor_name": user.get("name"),
        "product_name": prod.get("name"),
        "created_at": now_iso(),
        "created_by": user["id"],
        "verified": False,
        "verified_at": None,
        "verified_by": None,
        "verified_by_name": None,
        **payload.model_dump(),
    }
    await db.vendor_pricelists.insert_one(doc)
    return clean(doc)


@router.post("/vendor-portal/pricelists/bulk")
async def bulk_upload_pricelists(file: UploadFile = File(...), user=Depends(get_current_active_user)):
    """Vendor bulk-upload harga per SKU via CSV/XLSX.

    Expected columns: product_code, price, currency (optional), min_qty (optional),
    valid_from (opt), valid_until (opt), notes (opt).
    """
    vid = _require_vendor(user)
    db = get_db()
    content = await file.read()
    rows: list[dict] = []
    ext = (file.filename or "").lower()
    try:
        if ext.endswith(".xlsx") or ext.endswith(".xls"):
            from openpyxl import load_workbook
            import io as _io
            wb = load_workbook(_io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not any(r):
                    continue
                rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
        else:
            import csv as _csv, io as _io
            reader = _csv.DictReader(_io.StringIO(content.decode("utf-8-sig")))
            rows = [{k.strip().lower(): v for k, v in row.items() if k} for row in reader]
    except Exception as e:
        raise HTTPException(400, f"Gagal parsing file: {e}")
    prods_by_code = {p.get("code"): p async for p in db.products.find({}, {"_id": 0, "id": 1, "code": 1, "name": 1})}
    prods_by_id = {p["id"]: p for p in prods_by_code.values()}
    created = 0
    errors: list[dict] = []
    for i, r in enumerate(rows, start=2):
        code = str(r.get("product_code") or "").strip()
        pid = str(r.get("product_id") or "").strip()
        price_raw = r.get("price")
        prod = prods_by_id.get(pid) if pid else prods_by_code.get(code)
        if not prod:
            errors.append({"row": i, "error": f"produk tidak ditemukan (code={code}, id={pid})"})
            continue
        try:
            price = float(price_raw)
        except Exception:
            errors.append({"row": i, "error": f"harga tidak valid: {price_raw}"})
            continue
        doc = {
            "id": new_id(),
            "vendor_id": vid,
            "vendor_name": user.get("name"),
            "product_id": prod["id"],
            "product_name": prod.get("name"),
            "product_code": prod.get("code"),
            "price": price,
            "currency": (str(r.get("currency") or "IDR").upper()).strip() or "IDR",
            "min_qty": float(r.get("min_qty") or 1),
            "valid_from": str(r.get("valid_from") or "") or None,
            "valid_until": str(r.get("valid_until") or "") or None,
            "notes": str(r.get("notes") or "") or None,
            "verified": False,
            "created_at": now_iso(),
            "created_by": user["id"],
        }
        await db.vendor_pricelists.insert_one(doc)
        created += 1
    return {"ok": True, "created": created, "errors": errors, "total_rows": len(rows)}


@router.delete("/vendor-portal/pricelists/{plid}")
async def delete_vendor_pricelist(plid: str, user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    r = await db.vendor_pricelists.delete_one({"id": plid, "vendor_id": vid})
    if r.deleted_count == 0:
        raise HTTPException(404, "Pricelist tidak ditemukan atau bukan milik Anda")
    return {"ok": True}


@router.post("/pricelists/{plid}/verify")
async def verify_pricelist(plid: str, user=Depends(get_current_active_user)):
    """Procurement / admin verifies a vendor pricelist entry (toggles)."""
    if user["role"] not in ("admin", "procurement"):
        raise HTTPException(403, "Not allowed")
    db = get_db()
    pl = await db.vendor_pricelists.find_one({"id": plid}, {"_id": 0})
    if not pl:
        raise HTTPException(404, "Pricelist tidak ditemukan")
    new_state = not bool(pl.get("verified"))
    await db.vendor_pricelists.update_one({"id": plid}, {"$set": {
        "verified": new_state,
        "verified_at": now_iso() if new_state else None,
        "verified_by": user["id"] if new_state else None,
        "verified_by_name": user.get("name") if new_state else None,
    }})
    return {"ok": True, "verified": new_state}


@router.get("/vendor-portal/tenders/{tid}/price-suggestions")
async def tender_price_suggestions(tid: str, user=Depends(get_current_active_user)):
    """Historical PO price stats per tender item — helps vendor propose fair-range bid."""
    vid = _require_vendor(user)
    db = get_db()
    t = await db.tenders.find_one({"id": tid}, {"_id": 0})
    if not t:
        raise HTTPException(404, "Tender not found")
    invited = t.get("invited_vendor_ids") or []
    if invited and vid not in invited and not any(b.get("vendor_id") == vid for b in (t.get("bids") or [])):
        raise HTTPException(403, "Tender bukan untuk Anda")
    out: Dict[str, Any] = {}
    for it in (t.get("items") or []):
        pid = it.get("product_id")
        if not pid:
            continue
        prices: List[float] = []
        last_price: Optional[float] = None
        last_at: Optional[str] = None
        po_cursor = db.pos.find(
            {"items.product_id": pid, "status": {"$in": ["approved", "sent", "partial", "completed"]}},
            {"_id": 0, "items": 1, "created_at": 1, "po_number": 1},
        ).sort("created_at", -1).limit(50)
        async for po in po_cursor:
            for pit in (po.get("items") or []):
                if pit.get("product_id") == pid:
                    p = float(pit.get("price") or 0)
                    if p > 0:
                        prices.append(p)
                        if last_price is None:
                            last_price = p
                            last_at = po.get("created_at")
        if not prices:
            out[pid] = {"count": 0, "avg": None, "min": None, "max": None, "last": None, "last_at": None}
        else:
            out[pid] = {
                "count": len(prices),
                "avg": round(sum(prices) / len(prices), 2),
                "min": min(prices),
                "max": max(prices),
                "last": last_price,
                "last_at": last_at,
            }
    return {"tender_id": tid, "suggestions": out}


@router.post("/vendor-portal/tenders/{tid}/decline")
async def decline_tender(tid: str, user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    t = await db.tenders.find_one({"id": tid})
    if not t:
        raise HTTPException(404, "Not found")
    bids = t.get("bids", [])
    bids = [b for b in bids if b["vendor_id"] != vid]
    bids.append({
        "vendor_id": vid,
        "vendor_name": user["name"],
        "status": "declined",
        "submitted_at": now_iso(),
    })
    await db.tenders.update_one({"id": tid}, {"$set": {"bids": bids}})
    return {"ok": True}


@router.get("/vendor-portal/pos")
async def vendor_pos(user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    # Only fully approved / sent / completed POs (RFQs are in separate endpoint)
    q = {"vendor_id": vid, "status": {"$in": ["approved", "sent", "partial", "completed"]}}
    if user.get("is_pic"):
        q["assigned_pic_id"] = user["id"]
    return await db.pos.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.get("/vendor-portal/shipments")
async def vendor_shipments(user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    return await db.pos.find(
        {"vendor_id": vid, "shipping_status": {"$in": ["pending", "waiting_delivery", "partial"]}}, {"_id": 0}
    ).to_list(1000)


@router.get("/vendor-portal/invoices")
async def vendor_invoices(user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    return await db.invoices.find({"vendor_id": vid}, {"_id": 0}).sort("created_at", -1).to_list(1000)


async def _billing_status_for_po(db, pid: str) -> dict:
    """Compute per-item qty_billed / qty_remaining across all non-cancelled invoices for a PO."""
    po = await db.pos.find_one({"id": pid}, {"_id": 0})
    if not po:
        return None  # type: ignore
    items_out: list[dict] = []
    invoices = await db.invoices.find(
        {"po_id": pid, "status": {"$ne": "cancelled"}},
        {"_id": 0, "line_items": 1, "invoice_number": 1, "id": 1, "created_at": 1},
    ).to_list(500)
    billed_by_index: dict[int, float] = {}
    for inv in invoices:
        for li in (inv.get("line_items") or []):
            idx = int(li.get("po_item_index", -1))
            if idx >= 0:
                billed_by_index[idx] = billed_by_index.get(idx, 0.0) + float(li.get("qty_billed") or 0)
    for i, it in enumerate(po.get("items") or []):
        qo = float(it.get("qty") or 0)
        qb = billed_by_index.get(i, 0.0)
        items_out.append({
            "item_index": i,
            "product_id": it.get("product_id"),
            "product_name": it.get("product_name"),
            "product_code": it.get("product_code"),
            "unit": it.get("unit"),
            "qty_ordered": qo,
            "qty_billed": qb,
            "qty_remaining": max(0.0, qo - qb),
            "price": it.get("price"),
            "subtotal": it.get("subtotal"),
        })
    return {
        "po_id": pid,
        "po_number": po.get("po_number"),
        "vendor_id": po.get("vendor_id"),
        "currency": po.get("currency") or "IDR",
        "items": items_out,
        "invoice_count": len(invoices),
    }


@router.get("/vendor-portal/pos/{pid}/billing-status")
async def vendor_billing_status(pid: str, user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    po = await db.pos.find_one({"id": pid, "vendor_id": vid}, {"_id": 0, "id": 1})
    if not po:
        raise HTTPException(404, "PO tidak ditemukan atau bukan milik Anda")
    result = await _billing_status_for_po(db, pid)
    return result


@router.get("/pos/{pid}/billing-status")
async def admin_billing_status(pid: str, user=Depends(get_current_active_user)):
    db = get_db()
    result = await _billing_status_for_po(db, pid)
    if result is None:
        raise HTTPException(404, "PO tidak ditemukan")
    return result


class InvoiceLineItemIn(BaseModel):
    po_item_index: int
    qty_billed: float
    unit_price: Optional[float] = None
    discount_amount: Optional[float] = 0
    notes: Optional[str] = None


class InvoiceAttachmentIn(BaseModel):
    url: str
    filename: str
    size: Optional[int] = None
    content_type: Optional[str] = None
    kind: Optional[str] = "supporting"  # 'faktur_pajak' | 'bast' | 'supporting'


class InvoiceIn(BaseModel):
    po_id: str
    amount: float
    due_date: Optional[str] = None
    notes: Optional[str] = None
    ls_document_id: Optional[str] = None
    line_items: List[InvoiceLineItemIn] = []
    faktur_pajak_url: Optional[str] = None
    faktur_pajak_filename: Optional[str] = None
    faktur_pajak_number: Optional[str] = None
    bast_url: Optional[str] = None
    bast_filename: Optional[str] = None
    bast_number: Optional[str] = None
    attachments: List[InvoiceAttachmentIn] = []


@router.post("/vendor-portal/invoices")
async def submit_invoice(payload: InvoiceIn, user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    po = await db.pos.find_one({"id": payload.po_id})
    if not po:
        raise HTTPException(404, "PO not found")
    if po.get("vendor_id") != vid:
        raise HTTPException(403, "PO bukan milik Anda")
    # Faktur Pajak & BAST wajib
    if not payload.faktur_pajak_url:
        raise HTTPException(400, "Faktur Pajak wajib diupload")
    if not payload.bast_url:
        raise HTTPException(400, "BAST (Berita Acara Serah Terima) wajib diupload")
    if not payload.line_items:
        raise HTTPException(400, "Pilih minimal satu item PO yang akan ditagihkan")
    # Validate qty against remaining
    billing = await _billing_status_for_po(db, payload.po_id)
    remaining_by_idx = {it["item_index"]: it["qty_remaining"] for it in (billing.get("items") or [])}
    po_items = po.get("items") or []
    line_snapshots: list[dict] = []
    computed_amount = 0.0
    for li in payload.line_items:
        idx = li.po_item_index
        if idx < 0 or idx >= len(po_items):
            raise HTTPException(400, f"Item index {idx} tidak valid")
        rem = float(remaining_by_idx.get(idx, 0) or 0)
        if li.qty_billed <= 0:
            raise HTTPException(400, f"Qty untuk item {idx} harus > 0")
        if li.qty_billed > rem + 1e-6:
            raise HTTPException(400, f"Qty tagih {li.qty_billed} melebihi sisa {rem} pada item '{po_items[idx].get('product_name')}'")
        po_it = po_items[idx]
        unit_price = float(li.unit_price if li.unit_price is not None else po_it.get("price") or 0)
        disc = float(li.discount_amount or 0)
        subtotal = max(0.0, unit_price * li.qty_billed - disc)
        computed_amount += subtotal
        line_snapshots.append({
            "po_item_index": idx,
            "product_id": po_it.get("product_id"),
            "product_name": po_it.get("product_name"),
            "product_code": po_it.get("product_code"),
            "qty_billed": li.qty_billed,
            "unit_price": unit_price,
            "discount_amount": disc,
            "subtotal": subtotal,
            "notes": li.notes,
        })
    doc = {
        "id": new_id(),
        "invoice_number": gen_number("INV"),
        "vendor_id": vid,
        "vendor_name": po.get("vendor_name"),
        "po_id": payload.po_id,
        "po_number": po.get("po_number"),
        "amount": computed_amount,  # always trust server-computed to prevent client tampering
        "declared_amount": payload.amount,
        "computed_amount": computed_amount,
        "status": "outstanding",
        "due_date": payload.due_date,
        "notes": payload.notes,
        "ls_document_id": payload.ls_document_id,
        "is_bonded": po.get("po_type") == "BONDED",
        "line_items": line_snapshots,
        "items": po.get("items") or [],  # keep PO snapshot for legacy detail view
        "faktur_pajak_url": payload.faktur_pajak_url,
        "faktur_pajak_filename": payload.faktur_pajak_filename,
        "faktur_pajak_number": payload.faktur_pajak_number,
        "bast_url": payload.bast_url,
        "bast_filename": payload.bast_filename,
        "bast_number": payload.bast_number,
        "attachments": [a.model_dump() for a in payload.attachments],
        "untaxed_amount": po.get("untaxed_amount") or po.get("total") or 0,
        "tax_breakdown": po.get("tax_breakdown") or [],
        "taxes_snapshot": po.get("taxes_snapshot") or [],
        "amount_tax": po.get("amount_tax") or 0,
        "amount_total": po.get("amount_total") or po.get("total") or 0,
        "currency": po.get("currency") or "IDR",
        "exchange_rate": po.get("exchange_rate") or 1.0,
        "vendor_reply": po.get("vendor_reply"),
        "created_at": now_iso(),
    }
    await db.invoices.insert_one(doc)
    # Determine PO invoice status: if all items fully billed → 'complete', else 'partial'
    fresh_billing = await _billing_status_for_po(db, payload.po_id)
    remaining = sum(it.get("qty_remaining", 0) for it in (fresh_billing.get("items") or []))
    inv_status = "complete" if remaining <= 1e-6 else "partial"
    await db.pos.update_one({"id": payload.po_id}, {"$set": {"invoice_status": inv_status}})
    return clean(doc)


@router.get("/vendor-portal/invoices/{iid}")
async def vendor_invoice_detail(iid: str, user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    inv = await db.invoices.find_one({"id": iid, "vendor_id": vid}, {"_id": 0})
    if not inv:
        raise HTTPException(404, "Invoice tidak ditemukan")
    return inv


@router.get("/invoices/{iid}")
async def invoice_detail(iid: str, user=Depends(get_current_active_user)):
    db = get_db()
    inv = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not inv:
        raise HTTPException(404, "Invoice tidak ditemukan")
    if user["role"] == "vendor" and inv.get("vendor_id") != user.get("vendor_id"):
        raise HTTPException(403, "Not allowed")
    return inv


# ---------- LS documents (customs) ----------
class LSDocIn(BaseModel):
    doc_type: str = "LS"  # LS, PIB, BC23, etc
    po_id: Optional[str] = None
    reference_number: str
    hs_codes: List[str] = []
    file_url: Optional[str] = None
    notes: Optional[str] = None


@router.get("/vendor-portal/ls-documents")
async def list_ls_docs(user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    return await db.ls_documents.find({"vendor_id": vid}, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.post("/vendor-portal/ls-documents")
async def create_ls_doc(payload: LSDocIn, user=Depends(get_current_active_user)):
    vid = _require_vendor(user)
    db = get_db()
    doc = {
        "id": new_id(),
        "vendor_id": vid,
        "status": "submitted",
        "created_at": now_iso(),
        **payload.model_dump(),
    }
    await db.ls_documents.insert_one(doc)
    return clean(doc)


# Procurement side: list all LS docs
@router.get("/ls-documents")
async def list_all_ls(user=Depends(get_current_active_user)):
    db = get_db()
    return await db.ls_documents.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


# All invoices (procurement/finance side)
@router.get("/invoices")
async def list_all_invoices(user=Depends(get_current_active_user)):
    db = get_db()
    return await db.invoices.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.post("/invoices/{iid}/pay")
async def pay_invoice(iid: str, user=Depends(get_current_active_user)):
    if user["role"] not in ("admin", "finance"):
        raise HTTPException(403, "Not allowed")
    db = get_db()
    await db.invoices.update_one({"id": iid}, {"$set": {"status": "paid", "paid_at": now_iso()}})
    return {"ok": True}
